from random import randrange
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.support.ui import Select
import time
import openpyxl as xls

# Timing app reports contain headers, if your do not contain headers change the below number to 1
startrow = 2

# Location of the timing app report in excel format
path = "~/Downloads/test.xlsx"

wb = xls.load_workbook(path)
sheet = wb.active
row_tot = sheet.max_row
stdate = (sheet.cell(row=startrow, column = 1).value)
startdate = stdate.strftime("%m/%d/%Y")

# The code assumes that the first date of the week available for input in SpringAhead is saturday
targetdate = {"Saturday":4,"Sunday":5,"Monday":6,"Tuesday":7,"Wednesday":8,"Thursday":9,"Friday":10}

# Depending on the company you need to either round up to the nearest 15 or 30 minutes
# The items below randomizes between 15 and 30, most of the times it does not make a difference 
nr = {0:15,1:30}
r = randrange(2)

# Between the quotes, enter your credentials. Make sure there are no extra spaces.
username = " " #User name it is usually "First Last"
password = " " #password
uid = " " # Your user number, this can change with each contract, but usually remains the same, with the same contracting company.  This number can be obtained using Chrome developer tools after logging into springahead manually if not provided 
timerow = 1
descrow = 2
weektarget = sheet.cell(row=startrow, column = 1).value

# This function writes the appropriate time according to the date and description
def writetask(day,hours,desc,trow,drow):
    if desc == None:
        desc = " "
    browser.find_element_by_xpath("/html/body/div[2]/form[2]/div[1]/table/tbody[1]/tr["+str(trow)+"]/td["+str(day)+"]/input").send_keys(hours)
    browser.find_element_by_xpath("/html/body/div[2]/form[2]/div[1]/table/tbody[1]/tr["+str(drow)+"]/td[3]/input[1]").send_keys(desc)
    browser.find_element_by_class_name("addRow").click()

# Round up to the nearest 15/30 minutes 
def ceil_dt(dt, delta):
	return dt + (datetime.min - dt) % delta

browser = webdriver.Chrome('/YOUR/CHROME/DRIVER/LOCATION/')
browser.set_page_load_timeout(30)

# Depending on your timesheet type, the URL can be one of the two below
# browser.get("https://my.springahead.com/go/Account/Logon/REPLACE_WITH_COMPANY'S_NAME/")
browser.get("https://REPLACE_WITH_COMPANY'S_NAME.springahead.com/go/")

# Log into SpringAhead and select the correct week range based on the date of the first date cell in the excel spreadsheet exported from timing app
userID = browser.find_elements_by_id("UserName")
userID[0].send_keys(username)
browser.find_element_by_id("Password").send_keys(password)
browser.find_element_by_class_name("submit").click()
browser.get("https://REPLACE_WITH_COMPANY'S_NAME.springahead.com/vt/go?Timecard&tokenid=vte&userid="+uid+"&startDate="+startdate)

# Loops thru the each row in the spreadsheet and extracts particluar columns to use.
# The spreadsheet should only contain data for a single week. Spreadsheet spanding more than a single week will cause an error.
for i in range(row_tot-1):
    duration = sheet.cell(row=startrow, column = 2).value
    t1 = duration.strftime("%-H:%-M:%S")
    t2 = datetime.strptime(t1,'%H:%M:%S')
    
    # After the time object is converted to a string is passed to the round function 
    tt = ceil_dt(t2,timedelta(minutes = nr[r]))
    
    # Test for rounded hours and remove 00 minutes if they exist 
    if tt.strftime("%M")=="00":
        task_time = tt.strftime("%-H")
    else:
	    task_time = tt.strftime("%-H:%-M")
    
    sttime = desc = sheet.cell(row=startrow, column = 4).value
    f_sttime = sttime.strftime("%-I:%M:%S %p")
    sptime = desc = sheet.cell(row=startrow, column = 5).value
    f_sptime = sptime.strftime("%-I:%M:%S %p")
    desc = sheet.cell(row=startrow, column = 6).value
    date = sheet.cell(row=startrow, column = 1).value
    dayname = date.strftime("%A")
    daycell = targetdate[dayname]
    writetask(daycell,task_time,desc,timerow,descrow)
    startrow    +=1
    timerow     +=3
    descrow     +=3
    time.sleep(1)

# Submit Timesheet
time.sleep(3)
browser.find_element_by_xpath("""//*[@id="submitall"]""").click()

# Log Off
time.sleep(3)
browser.find_element_by_xpath("""//*[@id="titlebar"]/table/tbody/tr[1]/td/table/tbody/tr/td[3]/table/tbody/tr[2]/td/span/a[1]""").click()

# Quit Browser
time.sleep(2)
browser.quit()
